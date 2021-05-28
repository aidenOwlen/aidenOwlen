# -*- coding:utf-8 -*-
from PyQt4.QtCore import * 
from PyQt4.QtGui import * 
from PyQt4.QtCore import QThread 
from PyQt4 import QtCore,QtGui 
from openpyxl import load_workbook
from auto import *

global file 


class matchMain(QThread):
    def __init__(self):
        QThread.__init__(self)
    def run(self):
        global RUNNING,Main
        RUNNING = True
        wb = load_workbook(file)
        x = wb.get_sheet_names()

        #Major match 
        Assign = wb.get_sheet_by_name(x[1])
        Cand = wb.get_sheet_by_name(x[2])

        myCand = list(range(1,101))
        myAssign = list(range(1,101))
        Main = []
        Candidose = []

        MainRest = []
        i = 0
        while i+1 < len(Assign["A"]):
            MainListo = []
            i+= 1
            Id = Assign["A"][i].value
            skill1 = Assign["D"][i].value
            skill2 = Assign["E"][i].value
            driver = Assign["I"][i].value
            major = Assign["C"][i].value
            major2 = Assign["G"][i].value 
            major3 = Assign["H"][i].value
            location = Assign["B"][i].value
            k = 0
            while k +1< len(Cand["A"]) :
                k += 1
                IdC = Cand["A"][k].value
                CandidateN = "Candidate" + str(IdC)
                skillC1 = Cand["N"][k].value
                skillC2 = Cand["O"][k].value
                skillC3 = Cand["P"][k].value
                skillC4 = Cand["Q"][k].value
                skillC5 = Cand["R"][k].value
                skillC6 = Cand["S"][k].value
                skillC7 = Cand["T"][k].value
                skillC8 = Cand["U"][k].value
                skillC9 = Cand["V"][k].value
                skillC10 = Cand["W"][k].value
                skillC11 = Cand["X"][k].value
                skillC12 = Cand["Y"][k].value
                skillC13 = Cand["Z"][k].value
                skillC14 = Cand["AA"][k].value
                skillC15 = Cand["AB"][k].value
                skillC16 = Cand["AC"][k].value
                skillC17 = Cand["AD"][k].value
                skillC18 = Cand["AE"][k].value
                skillC19 = Cand["AF"][k].value
                skillC20 = Cand["AG"][k].value
                skillC21 = Cand["AH"][k].value
                skillCC = [skillC1,skillC2,skillC3,skillC4,skillC5,skillC6,skillC7,skillC8,skillC9,skillC10,skillC11,skillC12,skillC13,skillC14,skillC15,skillC16,skillC17,skillC18,skillC19,skillC20,skillC21]

                driverC = Cand["M"][k].value
                majorC = Cand["E"][k].value
                locationC1 = Cand["F"][k].value
                locationC2 = Cand["G"][k].value
                locationC3 = Cand["H"][k].value
                majorC2 = Cand["I"][k].value
                majorC3 = Cand["J"][k].value 
                majorC4 = Cand["K"][k].value


                if driver == "Driver's License" and driverC == "No":
                    driverMatch = "NotMatch"
                else:
                    driverMatch = "Match"
                if majorC in major:
                    MajorMatch  = "Match"
                else:
                    MajorMatch = "NotMatch"
                if locationC1 in location:
                    LocationMatch = 1
                elif locationC2 in location:
                    LocationMatch = 0.7
                elif locationC3 in location:
                    LocationMatch = 0.5
                else:
                    LocationMatch = 0
                SkillMatch = 0
                for kyu in skillCC:
                    if str(kyu) in str(skill1) or str(kyu) in str(skill2):
                        SkillMatch += 1

                if majorC2 in major2:
                    BusinessMatch = 2.5
                elif majorC2 in major3:
                    BusinessMatch = 2
                elif majorC3 in major2:
                    BusinessMatch =2
                elif majorC3 in major3:
                    BusinessMatch = 1.7
                elif majorC4 in major2:
                    BusinessMatch = 1.7
                elif majorC4 in major3:
                    BusinessMatch = 1.4
                else:
                    BusinessMatch = 0

                LocationMatch /= 2
                appendMatch = LocationMatch + SkillMatch
                Listo = [CandidateN,driverMatch,MajorMatch,appendMatch,BusinessMatch,SkillMatch,LocationMatch]
                MainListo.append(Listo)

            jk = 0
            jk1 = 0
            newC = ""
            newC1 = ""
            for J in MainListo:
                if J[1] == "NotMatch" or J[0] in Candidose:
                    pass
                  
                else:

                
                    if J[2] == "Match" and J[5] != 0:
                        ck = J[3]
                        if ck >= jk:
                            jk = J[3]
                            newC = J[0]
                    else:
                        ck1 = J[4] + J[5] + J[6]
                        if ck1 >= jk1:
                            jk1 = ck1
                            newC1 = J[0]

            if newC != "":
                    theCandidate = str(Id)+":"+newC
                    newCJ = newC
            elif newC1 != "":
                    theCandidate = str(Id)+":"+newC1
                    newCJ = newC1
            else:
                    theCandidate = "NoMatch:NoMatch"
            Main.append(theCandidate)
            Candidose.append(newCJ)
            self.emit(SIGNAL("added"),[Id-1,theCandidate])
            if RUNNING == False:
                break

class myInterface(QtGui.QMainWindow,Ui_MainWindow):
    def retranslateUi(self,MainWindow):
        super(__class__,self).retranslateUi(MainWindow)
        header = self.table.horizontalHeader()
        header.setResizeMode(0, QtGui.QHeaderView.Stretch)
        header.setResizeMode(1, QtGui.QHeaderView.Stretch)
        self.table.setRowCount(0)
        self.import_2.clicked.connect(self.importr)
        self.progressBar.setValue(0)
        self.export_2.clicked.connect(self.export)
        app.aboutToQuit.connect(self.runFalse)
    def export(self):
        self.pathToDl = str(QFileDialog.getExistingDirectory(self,"Select Directory"))
        from openpyxl import Workbook

        book = Workbook()
        sheet = book.active

       
        sheet.cell(row=1, column=1).value = "Assignment"
        sheet.cell(row=1, column=2).value = "Candidate"
        kui = 1
        for i in Main:
            kui += 1
            ij = i.split(":")
            assi = ij[0]
            cad = ij[1]
            sheet.cell(row=kui, column=1).value = assi
            sheet.cell(row=kui, column=2).value = cad



            book.save(self.pathToDl + '/OutPutExcel.xlsx')

    def runFalse(self):
        global RUNNING
        RUNNING = False
        
    def importr(self):
        global file 
        file = QFileDialog.getOpenFileName()
        RUNNING = False
        self.match = matchMain()
        self.match.start()
        self.connect(self.match,SIGNAL("added"),self.add)
    def add(self,listt):
        listto = listt[1].split(":")
        listt1 = listto[0]
        listt2 = listto[1]
        
        self.table.insertRow(listt[0])
        self.table.setItem(listt[0],0,QTableWidgetItem(str(listt1)))
        self.table.setItem(listt[0],1,QTableWidgetItem(str(listt2)))
        self.progressBar.setValue(listt[0]+1)





if __name__ == "__main__":
    import sys 
    app = QtGui.QApplication(sys.argv)
   
    
    MainWindow = QtGui.QMainWindow()
    ui = myInterface() 
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

